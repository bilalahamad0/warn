"""Tests for the Massachusetts (MA) WARN source."""

import re
from pathlib import Path

import pytest
from openpyxl import Workbook

import warn_sources
from warn_sources import ma as ma_module

# 17 real rows from the consolidated raw CSV the live fetch wrote on
# 2026-07-22 (FY22-FY27 workbooks + weekly tracker), covering both
# workbook generations and every cleaning quirk the module handles.
FIXTURE = Path(__file__).parent / "fixtures" / "ma_consolidated_sample.csv"

ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_registry_contains_massachusetts():
    assert "ma" in warn_sources.SOURCES
    cls = warn_sources.SOURCES["ma"]
    assert cls.code == "ma"
    assert cls.name == "Massachusetts"
    assert cls.source_url.startswith("https://www.mass.gov/")


def test_paths_use_per_state_layout(tmp_path):
    src = warn_sources.get_source("ma", tmp_path)
    assert src.paths.root == tmp_path / "states" / "ma"
    assert src.paths.latest.name == "warn_latest.json"
    assert src.paths.raw.name == "raw_download.csv"


# ---------------------------------------------------------------------------
# Offline parse against the real-data fixture
# ---------------------------------------------------------------------------


@pytest.fixture
def parsed(tmp_path):
    src = warn_sources.get_source("ma", tmp_path)
    return src.parse(FIXTURE)


def test_parse_columns(parsed):
    assert list(parsed.columns) == [
        "company", "notice_date", "effective_date", "employees",
        "county", "city",
    ]


def test_parse_keeps_all_company_rows(parsed):
    assert len(parsed) == 17
    assert (parsed["company"].str.strip() != "").all()


def test_parse_dates_are_iso_or_none(parsed):
    for col in ("notice_date", "effective_date"):
        for val in parsed[col]:
            assert val is None or ISO_DATE.match(val), (col, val)


def test_parse_employees_are_int(parsed):
    assert all(isinstance(v, int) for v in parsed["employees"])


def test_received_range_takes_first_date(parsed):
    # Gen-1 quirk: "07/07/2021 - (08/30/2021)" in the received column.
    sodexo = parsed[parsed["company"] == "Sodexo @ Suffolk University"]
    assert sodexo["notice_date"].tolist() == ["2021-07-07"]
    assert sodexo["employees"].tolist() == [74]


def test_layoff_range_takes_first_date(parsed):
    merck = parsed[parsed["company"].str.startswith("Merck & Co.")]
    assert merck["effective_date"].tolist() == ["2022-05-31"]  # 2-digit yrs
    coke = parsed[parsed["company"] == "The Coca-Cola Company"]
    assert coke["effective_date"].tolist() == ["2026-08-15"]  # "&"-list


def test_in_ma_figure_preferred_over_total(parsed):
    akebia = parsed[parsed["company"].str.startswith("Akebia")]
    assert akebia["employees"].tolist() == [91]  # "181 (total) 91 in MA"
    aspiration = parsed[parsed["company"] == "Aspiration Partners, Inc."]
    assert aspiration["employees"].tolist() == [1]  # "180 (1 resides in MA)"


def test_range_and_correction_counts(parsed):
    sqz = parsed[parsed["company"] == "SQZ Biotechnologies Company"]
    assert sqz["employees"].tolist() == [70]  # "70 - 80" lower bound
    takeda = parsed[parsed["company"].str.startswith("Takeda")]
    assert takeda["employees"].tolist() == [138]  # "Up to 138"
    # And its prose layoff cell has no full date -> None, never a
    # synthesized day.
    assert takeda["effective_date"].tolist() == [None]
    immunity = parsed[parsed["company"] == "Immunity Bio. Inc."]
    assert immunity["employees"].tolist() == [1]  # "min. 1 - max. 10"


def test_no_count_cells_map_to_zero(parsed):
    bridal = parsed[parsed["company"] == "David's Bridal"]
    assert bridal["employees"].tolist() == [0]  # "work in progress"
    emd = parsed[parsed["company"] == "EMD Serono, Inc,"]
    assert emd["employees"].tolist() == [0]  # "t/b/d"


def test_date_typo_correction(parsed):
    edaron = parsed[parsed["company"].str.startswith("Edaron")]
    assert edaron["effective_date"].tolist() == ["2025-03-31"]  # 3/31/12025


def test_updated_row_blank_received_is_none(parsed):
    chl = parsed[parsed["company"].str.startswith("*Updated*")]
    assert chl["notice_date"].tolist() == [None]  # blank RECEIVED cell
    assert chl["employees"].tolist() == [131]
    # Company kept exactly as published, marker included.
    assert chl["company"].tolist() == [
        "*Updated* Community Healthlink, Inc. (aka CHL)"
    ]


def test_city_strips_trailing_ma_suffix(parsed):
    laboure = parsed[parsed["company"].str.startswith("Labour")]
    assert laboure["city"].tolist() == ["Milton"]  # was "Milton, MA"
    assert laboure["notice_date"].tolist() == ["2026-07-01"]
    chl = parsed[parsed["company"].str.startswith("*Updated*")]
    assert chl["city"].tolist() == ["Leominster, Webster & Worcester"]


def test_region_promoted_to_county(parsed):
    assert set(parsed["county"]) <= {
        "Boston", "Central", "Metro Southwest", "Northeast", "Southeast",
        "West", "Western", "Remote", "Remote/National", "Statewide",
    }


def test_unify_produces_unified_schema(tmp_path, parsed):
    src = warn_sources.get_source("ma", tmp_path)
    df = src.unify(parsed)
    assert list(df.columns)[: len(warn_sources.UNIFIED_FIELDS)] == (
        warn_sources.UNIFIED_FIELDS
    )
    assert (df["state"] == "MA").all()
    # MA publishes no layoff type, address, or industry — never
    # fabricated.
    assert (df["layoff_type"] == "").all()
    assert (df["address"] == "").all()
    assert (df["industry"] == "").all()


# ---------------------------------------------------------------------------
# Workbook-grid extraction (both sheet generations)
# ---------------------------------------------------------------------------


def _grid(ws):
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_grid_rows_gen1_region_sheet():
    # FY22 "Central"-style sheet: title row, header row missing its
    # "Date Received" cell (received falls back to column 0), region
    # from the sheet name.
    wb = Workbook()
    ws = wb.active
    ws.append([None, "2022 WARN Report - Central", None, None, None])
    ws.append([None, None, None, None, None])
    ws.append([None, "Company Name", "City", "Layoff Date", "# Affected"])
    ws.append(["10/8/2021", "Valdilvia Logistics LLC", "Milford",
               "10/16/2021", 52])
    rows = ma_module._grid_rows(_grid(ws), default_region="Central")
    assert rows == [{
        "received": "10/8/2021",
        "employer": "Valdilvia Logistics LLC",
        "city": "Milford",
        "region": "Central",
        "layoff_dates": "10/16/2021",
        "employees": "52",
    }]


def test_grid_rows_gen2_single_sheet():
    # FY24+ layout: REGION column wins over any sheet-name default, and
    # datetime cells are ISO-stamped at consolidation time.
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.append(["RECEIVED", "EMPLOYER", "CITY/TOWN", "REGION",
               "DATE(S) OF LAYOFFS", "# EMPLOYEES IMPACTED"])
    ws.append([datetime(2026, 6, 15), "The Coca-Cola Company",
               "Northampton", "Western", "8/15/26 & 11/30/26", 175])
    ws.append([None, None, None, None, None, None])  # spacer -> dropped
    rows = ma_module._grid_rows(_grid(ws), default_region="WARN Report FY26")
    assert rows == [{
        "received": "2026-06-15",
        "employer": "The Coca-Cola Company",
        "city": "Northampton",
        "region": "Western",
        "layoff_dates": "8/15/26 & 11/30/26",
        "employees": "175",
    }]


def test_find_data_links_sorts_fy_oldest_first():
    # Anchor hrefs verbatim from the live landing page (2026-07-22).
    html = """
    <a href="https://www.mass.gov/files/csv/2026-07/WARN%20Report%20FY%20\
2027%20week%20ending%2007-17-2026.csv">Download table data as CSV</a>
    <a href="https://www.mass.gov/doc/fy26-warn-report-0/download">FY26</a>
    <a href="https://www.mass.gov/doc/fy25-warn-report/download">FY25</a>
    <a href="https://www.mass.gov/doc/fy22-warn-report/download">FY22</a>
    <a href="/how-to/submit-a-warn-notice">Submit a WARN notice</a>
    """
    xlsx, csvs = ma_module._find_data_links(html)
    assert [u.split("/doc/")[1] for u in xlsx] == [
        "fy22-warn-report/download",
        "fy25-warn-report/download",
        "fy26-warn-report-0/download",
    ]
    assert len(csvs) == 1
    assert csvs[0].endswith("07-17-2026.csv")


def test_clean_date_rules():
    assert ma_module._clean_date("2026-06-30") == "2026-06-30"
    assert ma_module._clean_date("7/15/2026") == "2026-07-15"
    assert ma_module._clean_date("10/3/2022 - 2/4/2023") == "2022-10-03"
    assert ma_module._clean_date("T/B/D") is None
    assert ma_module._clean_date("Spring 2025 - Spring 2026") is None
    assert ma_module._clean_date("11/31/2021") == "2021-11-30"
    assert ma_module._clean_date("") is None
    assert ma_module._clean_date(None) is None
    assert ma_module._clean_date("1/1/1900") is None  # typo window
