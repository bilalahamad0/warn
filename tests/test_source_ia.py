"""Tests for the Iowa WARN source (warn_sources/ia.py).

Offline: the fixture workbooks are built with openpyxl from rows really
fetched from workforce.iowa.gov (and BLN's historic archive) on
2026-07-21, reproducing the feed's documented quirks: a title row above
the header, hand-typed date typos ("3/32/2026"), string and datetime
date cells mixed in one column, trailing whitespace everywhere, a
zero-employee amendment row, and the headerless "Sheet2" program-year
compilation that duplicates the yearly sheets and ends in a grand-total
footer.
"""

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

import warn_sources
from warn_sources import ia

ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")

_HEADER = [
    "Company", "Address Line 1", "City", "County", "St", "ZIP",
    "Notice Type", "Emp #", "Notice Date", "Layoff Date",
    "Local Workforce Area", "Industry",
]

# Real rows fetched from the live "Iowa WARN Log" workbook on 2026-07-21.
_CURRENT_ROWS = [
    ["CNH Industrial America LLC ", "1930 Des Moines Ave", "Burlington",
     "Des Moines", "IA", 52601, "Closing ", 7,
     datetime(2026, 1, 20), "3/32/2026", "Mississippi Valley ",
     "Manufacturing "],
    ["John Deere Ottumwa Works ", "928 E Vine St", "Ottumwa", "Wapello",
     "IA ", 52501, "Mass Layoff", 75, "01/06/2025", datetime(2025, 2, 7),
     "South Central Iowa ", "Manufacturing"],
    ["River Hills Community Center", "100 West Main St", "Richland ",
     "Keokuk", "IA ", 52585, "Closing ", 5, datetime(2026, 7, 17),
     datetime(2026, 7, 31), "Eastern Iowa ",
     "Health Care and Social Assistance"],
    ["Wells Fargo ", "800 S Jordan Creek Pkwy", "West Des Moines", "Polk",
     "IA", 50266, "Amendment - Additional Employees", 21, "12/10/2024",
     datetime(2025, 2, 10), "Central Iowa", "Finance and Insurance "],
    ["United States Cellular Corporation", "515 49th Ave. Dr. SW",
     "Cedar Rapids", "Linn", "IA ", 52404, "Amendment - Change in Date", 0,
     datetime(2025, 6, 11), datetime(2025, 7, 1), "East Central Iowa ",
     "Information "],
]

# Real rows from BLN's 2011-2018 historic archive workbook (10 columns:
# no Local Workforce Area / Industry, header on the first row).
_HISTORIC_ROWS = [
    ["United HR Direct", "3022 Airport Blvd.", "Waterloo", "Black Hawk",
     "IA", "50704", "Mass Layoff", 52, datetime(2011, 1, 13),
     datetime(2011, 3, 8)],
    ["Electrolux", "400 Des Moines Street", "Webster City", "Hamilton",
     "IA", "50595", "Amendment", 510, datetime(2011, 1, 27),
     datetime(2011, 3, 31)],
]


def _build_current(path: Path) -> Path:
    """Current-log fixture: title row + header + rows, plus 'Sheet2'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2026"
    ws.append(["Iowa WARN Log", "Updated: 07-17-2026"])
    ws.append(_HEADER)
    for row in _CURRENT_ROWS:
        ws.append(row)
    # The duplicate program-year compilation: no header row, repeats a
    # notice already in a yearly sheet, ends in a grand-total footer.
    s2 = wb.create_sheet("Sheet2")
    s2.append([None] * 12)
    s2.append(["Smithfield Packaged Meats Corp", "612 Adventureland Dr, NE",
               "Altoona", "Polk", "IA", 50009, "Closing ", 319,
               datetime(2024, 7, 2), datetime(2024, 8, 30),
               "Central Iowa ", "Manufacturing"])
    s2.append([None] * 7 + [7632] + [None] * 4)
    wb.save(path)
    return path


def _build_historic(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "WARN Log 7_12_17- 2"
    ws.append(_HEADER[:6] + ["Notice Type", "Emp #", "Notice Date",
                             "Layoff Date"])
    for row in _HISTORIC_ROWS:
        ws.append(row)
    wb.save(path)
    return path


def _parse_fixtures(tmp_path):
    """Consolidate both fixture workbooks and parse, like fetch() would."""
    rows = ia._extract_workbook_rows(_build_current(tmp_path / "cur.xlsx"))
    rows += ia._extract_workbook_rows(_build_historic(tmp_path / "hist.xlsx"))
    raw = tmp_path / "raw_download"
    raw.write_text(json.dumps(rows, default=str))
    return warn_sources.get_source("ia", tmp_path).parse(raw)


def test_registry_contains_iowa():
    assert "ia" in warn_sources.SOURCES
    cls = warn_sources.SOURCES["ia"]
    assert cls.enabled
    assert cls.name == "Iowa"


def test_parse_columns_and_types(tmp_path):
    df = _parse_fixtures(tmp_path)
    assert list(df.columns) == [
        "company",
        "notice_date",
        "effective_date",
        "employees",
        "layoff_type",
        "county",
        "city",
        "address",
        "industry",
    ]
    assert str(df["employees"].dtype) == "int64"
    for col in ("notice_date", "effective_date"):
        for val in df[col]:
            assert val is None or ISO.match(val), f"bad {col}: {val!r}"


def test_parse_skips_headerless_duplicate_sheet(tmp_path):
    df = _parse_fixtures(tmp_path)
    # 5 current rows + 2 historic rows; Sheet2's duplicate notice and its
    # grand-total footer must not leak in.
    assert len(df) == 7
    assert "Smithfield Packaged Meats Corp" not in df["company"].tolist()


def test_parse_applies_bln_date_corrections(tmp_path):
    df = _parse_fixtures(tmp_path)
    cnh = df[df.company == "CNH Industrial America LLC"].iloc[0]
    # Hand-typed "3/32/2026" layoff date -> 2026-03-23 (BLN correction);
    # the notice date is never copied into it.
    assert cnh["effective_date"] == "2026-03-23"
    assert cnh["notice_date"] == "2026-01-20"


def test_parse_handles_mixed_string_and_datetime_dates(tmp_path):
    df = _parse_fixtures(tmp_path)
    deere = df[df.company == "John Deere Ottumwa Works"].iloc[0]
    assert deere["notice_date"] == "2025-01-06"    # "01/06/2025" string
    assert deere["effective_date"] == "2025-02-07"  # datetime cell


def test_parse_strips_whitespace_and_maps_fields(tmp_path):
    df = _parse_fixtures(tmp_path)
    rh = df[df.company == "River Hills Community Center"].iloc[0]
    assert rh["city"] == "Richland"
    assert rh["county"] == "Keokuk"
    assert rh["address"] == "100 West Main St"
    assert rh["industry"] == "Health Care and Social Assistance"
    assert rh["layoff_type"] == "Closing"


def test_parse_keeps_amendment_notice_types(tmp_path):
    df = _parse_fixtures(tmp_path)
    wf = df[df.company == "Wells Fargo"].iloc[0]
    assert wf["layoff_type"] == "Amendment - Additional Employees"
    assert wf["employees"] == 21


def test_parse_zero_employee_rows_stay_int(tmp_path):
    df = _parse_fixtures(tmp_path)
    usc = df[df.company == "United States Cellular Corporation"].iloc[0]
    assert usc["employees"] == 0


def test_parse_historic_backfill_rows(tmp_path):
    df = _parse_fixtures(tmp_path)
    uhd = df[df.company == "United HR Direct"].iloc[0]
    assert uhd["notice_date"] == "2011-01-13"
    assert uhd["effective_date"] == "2011-03-08"
    assert uhd["employees"] == 52
    assert uhd["industry"] == ""   # historic workbook publishes none
    lux = df[df.company == "Electrolux"].iloc[0]
    assert lux["layoff_type"] == "Amendment"
