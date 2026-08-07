"""
warn_sources.ky
---------------
Kentucky — WARN notices published by the Kentucky Career Center
(kyworks.ky.gov, "Rapid Response Layoffs and Closures").

The state posts two workbooks behind ``a.btn-block`` buttons on the host
page (link discovery vendored from Big Local News' Apache-2.0 warn-scraper,
warn/scrapers/ky.py):

* a dated current report ("WARN Notice Report", XLSX; the filename changes
  every refresh, and the state has flip-flopped between CSV/XLS/XLSX), and
* a "Prior Year Warn Notices" archive workbook, one sheet per year
  (2017-present), which provides the backfill.

DELIBERATE DIVERGENCE FROM BLN: Big Local News' warn-transformer
(warn_transformer/transformers/ky.py) maps Kentucky's dates SWAPPED —
notice_date from "Projected Date" and effective_date from "Date Received".
This module used to replicate that, but in the unified national schema —
where every other state's notice_date is the filing date — the swap put
future "notice dates" (projected layoffs up to months out) at the top of the
national dashboard's newest-first table, and 67% of KY records showed an
effective date before their notice date. So here the mapping follows the
column names' plain meaning instead: ``notice_date`` <- "Date Received" (the
filing date), ``effective_date`` <- "Projected Date" (the layoff date).
Neither date is ever copied into the other.

Other feed quirks handled here:
* the 2024/2025 archive sheets ship the County column with a *blank* header
  cell between "Region" and "Company Name" — recovered positionally;
* the 2019 sheet contains raw Excel date serials (43490 → 2019-01-25);
* the 2017 sheet's NAICS column was mangled into datetimes by Excel — such
  values are dropped rather than published as industry codes;
* the current report and the archive overlap for recent months, so the
  consolidated feed is de-duplicated on (company, dates, employees).
"""

import csv
import io
import json
import logging
import re
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import warn_monitor
from .base import Source

log = logging.getLogger("warn_sources")

HOSTPAGE = (
    "https://kyworks.ky.gov/Services/Pages/"
    "Rapid-Response-Layoffs-and-Closures.aspx"
)
BASEURL = "https://kyworks.ky.gov"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)

# Column crosswalk vendored from Big Local News' Apache-2.0 warn-scraper
# (warn/scrapers/ky.py) — maps every header KY has used to a stable field.
_CROSSWALK = {
    "Closure or Layoff?": "closure_or_layoff",
    "Company Name": "company",
    "Company: Company Name": "company",
    "County": "county",
    "Date Received": "date_received",
    "Employees": "employees",
    "NAICS": "NAICS",
    "NAICS Code": "NAICS",
    "Notice Link": "notice_url",
    "Notice Type": "source",
    "Notice URL": "notice_url",
    "Notice: Notice Number": "notice_number",
    "Number of Employees Affected": "employees",
    "Projected Date": "date_effective",
    "Region": "region",
    "Trade": "trade",
    "Type of Employees Affected": "union_affected",
    "Workforce Board": "region",
}

# Junk placeholders KY has used in date cells (from BLN date_corrections).
_JUNK_DATES = {"", "?", "n/a", "tbd", "unknown", "see warn", "november"}

_MAX_JOBS = 10000  # BLN sanity cap for KY employee counts
_MIN_YEAR = 1997   # BLN minimum_year for KY

_OUTPUT_COLUMNS = [
    "company",
    "notice_date",
    "effective_date",
    "employees",
    "layoff_type",
    "county",
    "industry",
]


def _clean_str(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


def _excel_serial_to_iso(num):
    """Convert an Excel date serial (e.g. 43490) to an ISO date string."""
    try:
        num = float(num)
    except (TypeError, ValueError):
        return None
    if not 30000 <= num <= 60000:  # ~1982..2064: plausible serial window
        return None
    return (date(1899, 12, 30) + timedelta(days=int(num))).isoformat()


def _validate_iso(iso):
    """Accept only real ISO dates within a sane year window, else None."""
    if not iso or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(iso)):
        return None
    year = int(str(iso)[:4])
    if not _MIN_YEAR <= year <= date.today().year + 10:
        return None
    return str(iso)


def _clean_date(val):
    """Raw feed date cell -> ISO YYYY-MM-DD string or None (never junk)."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return _validate_iso(_excel_serial_to_iso(val))
    s = str(val).strip()
    if not s or s.lower() in _JUNK_DATES:
        return None
    if re.fullmatch(r"\d{5}(\.0)?", s):  # serial that arrived as text
        return _validate_iso(_excel_serial_to_iso(s))
    return _validate_iso(warn_monitor._safe_date(s))


def _clean_employees(val) -> int:
    """Raw employees cell -> int (0 when the state publishes no count)."""
    if isinstance(val, str):
        # Range/suffix cleanup vendored from BLN warn-transformer ky.py
        # transform_jobs: "75 - 100" -> 75, "90 +/-" -> 90.
        s = val.split("-")[0]
        s = s.replace("+/-", "").replace("+/", "").replace("+", "")
        val = re.split(r" {5,}", s.strip())[0].strip()
    n = warn_monitor._safe_int(val)
    if n is None or n < 0 or n > _MAX_JOBS:
        return 0
    return n


def _clean_naics(val) -> str:
    """NAICS cell -> code string; drops the 2017 sheet's datetime garbage."""
    n = warn_monitor._safe_int(val)
    if n is None or not 10 <= n <= 999999:
        return ""
    return str(n)


def _clean_company(val) -> str:
    s = _clean_str(val)
    if s.lower() in {"company name", "company: company name"}:
        return ""  # stray header echo
    return s


def _grid_to_rows(grid):
    """One sheet's non-empty rows -> list of crosswalk-keyed dicts.

    Header/footer handling vendored from BLN warn-scraper ky.py start_xlsx:
    the header row is the one containing "Company Name"; a "Total ... Count"
    row (and everything after it) is a footer; repeated headers are skipped.
    """
    header_idx = None
    end_idx = None
    for i, row in enumerate(grid):
        joined = " ".join(str(c) for c in row if c is not None)
        if header_idx is None and "Company Name" in joined:
            header_idx = i
            continue
        if header_idx is not None and end_idx is None:
            if "Total" in joined and "Count" in joined:
                end_idx = i
    if header_idx is None:
        return []
    if end_idx is None:
        end_idx = len(grid)

    headers = []
    nulls = 0
    for cell in grid[header_idx]:
        name = _clean_str(cell)
        if not name:
            nulls += 1
            headers.append(f"null_{nulls}")
        else:
            headers.append(_CROSSWALK.get(name, name))
    # KY quirk: the 2024/2025 archive sheets ship the County column with a
    # blank header cell between Region and Company Name — recover it.
    if "county" not in headers and "region" in headers:
        i = headers.index("region") + 1
        if i < len(headers) and headers[i].startswith("null_"):
            headers[i] = "county"

    rows = []
    for row in grid[header_idx + 1:end_idx]:
        joined = " ".join(str(c) for c in row if c is not None)
        if "Company Name" in joined or joined.startswith("Date Received"):
            continue  # repeated header row
        rec = {}
        for field, val in zip(headers, row):
            if field.startswith("null_") or val is None:
                continue
            if isinstance(val, (datetime, date)):
                val = str(val)
            elif isinstance(val, str):
                val = val.strip()
            if val == "":
                continue
            rec[field] = val
        if rec:
            rows.append(rec)
    return rows


def _extract_workbook_rows(xlsx_source):
    """Every sheet of an XLSX (path or file-like) -> crosswalked dicts."""
    wb = load_workbook(filename=xlsx_source, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        grid = [
            list(r)
            for r in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() != "" for c in r)
        ]
        rows.extend(_grid_to_rows(grid))
    return rows


def _extract_csv_rows(text):
    """CSV variant of the feed (used briefly in late 2025) -> dicts."""
    grid = [
        row
        for row in csv.reader(text.splitlines())
        if any(str(c).strip() for c in row)
    ]
    return _grid_to_rows(grid)


def _consolidate(row_groups):
    """Union the current report with the archive, de-duplicated.

    The current report overlaps the archive's most recent sheet; the first
    occurrence wins (callers pass the richer current report first). County
    is excluded from the key because the two files spell it differently
    ("Jefferson" vs "Jefferson County - Louisville"); employees is included
    so same-day multi-notice filings (e.g. GDI 2024-01-04) survive.
    """
    seen = {}
    for rows in row_groups:
        for r in rows:
            key = (
                _clean_str(r.get("company")).lower(),
                _clean_date(r.get("date_received")),
                _clean_date(r.get("date_effective")),
                _clean_employees(r.get("employees")),
            )
            if key not in seen:
                seen[key] = r
    return list(seen.values())


class KentuckyCareerCenter(Source):
    code = "ky"
    name = "Kentucky"
    agency = "Kentucky Career Center (Education & Labor Cabinet)"
    source_url = HOSTPAGE
    cadence = "weekly"

    def _get(self, url, tries=3):
        """Polite GET: browser UA, 60s timeout, up to 3 backed-off tries."""
        last_err = None
        for attempt in range(tries):
            if attempt:
                time.sleep(2 * attempt)
            try:
                resp = requests.get(
                    url, headers={"User-Agent": USER_AGENT}, timeout=60
                )
                resp.raise_for_status()
                return resp
            except requests.RequestException as e:
                last_err = e
                log.warning(f"[KY] GET {url} failed (try {attempt + 1}): {e}")
        raise last_err

    def _discover(self, html):
        """Find the current-report and prior-year links on the host page.

        Vendored from BLN warn-scraper ky.py: the report is the a.btn-block
        button whose text contains "report" (last match wins).
        """
        soup = BeautifulSoup(html, "html.parser")
        current = None
        prior = None
        for a in soup.find_all("a", class_="btn-block"):
            text = a.get_text(" ", strip=True).lower()
            href = a.get("href") or ""
            if not href or not href.lower().endswith((".xlsx", ".csv")):
                continue
            if "report" in text:
                current = urljoin(BASEURL, href)
            elif "prior year" in text:
                prior = urljoin(BASEURL, href)
        if current is None:
            raise RuntimeError(
                "KY WARN report link not found on host page "
                "(layout changed or feed moved)"
            )
        if prior is None:
            log.warning("[KY] prior-year archive link not found; "
                        "using current report only")
        return current, prior

    def fetch(self, force: bool = False) -> tuple:
        self.paths.ensure()
        host = self._get(HOSTPAGE)
        current_url, prior_url = self._discover(host.text)

        groups = []
        for url in [u for u in (current_url, prior_url) if u]:
            time.sleep(1)  # politeness: max 1 request/second/host
            resp = self._get(url)
            if url.lower().endswith(".csv"):
                groups.append(_extract_csv_rows(resp.text))
            else:
                groups.append(_extract_workbook_rows(io.BytesIO(resp.content)))

        rows = _consolidate(groups)
        self.paths.raw.write_text(json.dumps(rows, indent=1, default=str))

        meta = warn_monitor._load_meta(self.paths.meta)
        new_hash = warn_monitor._file_hash(self.paths.raw)
        changed = force or new_hash != meta.get("file_hash", "")
        meta.update(
            {
                "file_hash": new_hash,
                "last_checked": datetime.now(timezone.utc).isoformat(),
                "url": current_url,
                "prior_url": prior_url,
                "row_count": len(rows),
            }
        )
        warn_monitor._save_meta(meta, self.paths.meta)
        return changed, str(self.paths.raw)

    def parse(self, raw_path) -> pd.DataFrame:
        rows = json.loads(Path(raw_path).read_text())
        out = []
        for r in rows:
            company = _clean_company(r.get("company"))
            if not company:
                continue
            out.append(
                {
                    "company": company,
                    # Plain-meaning mapping — deliberately NOT BLN's swapped
                    # field map; see the module docstring.
                    "notice_date": _clean_date(r.get("date_received")),
                    "effective_date": _clean_date(r.get("date_effective")),
                    "employees": _clean_employees(r.get("employees")),
                    "layoff_type": _clean_str(r.get("closure_or_layoff")),
                    "county": _clean_str(r.get("county")),
                    "industry": _clean_naics(r.get("NAICS")),
                }
            )
        return pd.DataFrame(out, columns=_OUTPUT_COLUMNS)
