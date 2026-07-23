"""
warn_sources.ia
---------------
Iowa — WARN notices published by Iowa Workforce Development on the
"Iowa WARN Log" workbook linked from the agency's WARN page.

Link discovery is vendored from Big Local News' Apache-2.0 warn-scraper
(warn/scrapers/ia.py): the host page anchor whose text mentions "Excel"
points at the current log. The workbook ships one sheet per year
(2021-present) plus an unlabelled program-year compilation sheet
("Sheet2", Jul 2024 - Jun 2025) that *duplicates* the yearly sheets,
carries no header row, and ends in a grand-total footer — sheets without
a recognisable header row are skipped so no notice is ever counted twice.

Backfill (also from BLN ia.py): BLN's public GCS bucket hosts an archived
Iowa log covering 2011 - May 2018. The state removed pre-2021 rows from
the live workbook, so the 2018-06 .. 2021-07 window is a gap in the
state's own record, not a scraping loss.

Field mapping vendored exactly from BLN's Apache-2.0 warn-transformer
(warn_transformer/transformers/ia.py): company <- "Company",
city <- "City", notice_date <- "Notice Date",
effective_date <- "Layoff Date", employees <- "Emp #" — neither date is
ever copied into the other. BLN's date_corrections for Iowa's hand-typed
date typos ("3/32/2026", "9/1/8/2020", "59/09/2025", …) are vendored
verbatim. Iowa additionally publishes County, Address Line 1, Notice
Type and Industry, carried into the unified county / address /
layoff_type / industry fields (the historic workbook lacks Industry).
"""

import io
import json
import logging
import re
import time
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import warn_monitor
from .base import Source

log = logging.getLogger("warn_sources")

HOSTPAGE = "https://workforce.iowa.gov/employers/business-resources/warn"
BASEURL = "https://workforce.iowa.gov"
HISTORIC_URL = (
    "https://storage.googleapis.com/bln-data-public/"
    "warn-layoffs/ia_historical_2018.xlsx"
)
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)

# Header -> unified field crosswalk (BLN warn-transformer ia.py fields dict,
# extended with the extra columns Iowa really publishes). Unmapped headers
# (St/State, ZIP, Local Workforce Area) are dropped.
_CROSSWALK = {
    "Company": "company",
    "Address Line 1": "address",
    "City": "city",
    "County": "county",
    "Notice Type": "layoff_type",
    "Emp #": "employees",
    "Notice Date": "notice_date",
    "Layoff Date": "effective_date",
    "Industry": "industry",
}

# Hand-typed date typos in the feed -> real dates. Vendored verbatim from
# BLN warn-transformer ia.py date_corrections (keys are the raw cell text
# exactly as str() renders it).
_DATE_CORRECTIONS = {
    "9/1/8/2020": "2020-01-08",
    "4/26/21": "2021-04-26",
    "2021-04-30 00:00:00": "2021-04-30",
    "7/14/21": "2021-07-14",
    "7/12/21": "2021-07-12",
    "2027-07-27 00:00:00": "2024-07-27",
    "2025-12-31 00:00:00": "2025-12-31",
    "59/09/2025": "2025-05-19",
    "3/32/2026": "2026-03-23",
    "4/30/206": "2026-04-30",
}

# BLN date_format for Iowa, plus %m/%d/%y for stray 2-digit years.
_DATE_FORMATS = ("%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%y")

_MAX_JOBS = 10000  # BLN default sanity cap for employee counts
_MIN_YEAR = 2000   # historic log starts 2011; anything earlier is a typo

_OUTPUT_COLUMNS = [
    "company",
    "notice_date",
    "effective_date",
    "employees",
    "layoff_type",
    "county",
    "city",
    "address",
    "industry",
]


def _clean_str(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


def _validate_iso(iso):
    """Accept only real ISO dates within a sane year window, else None."""
    if not iso or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(iso)):
        return None
    year = int(str(iso)[:4])
    if not _MIN_YEAR <= year <= date.today().year + 10:
        return None
    return str(iso)


def _clean_date(val):
    """Raw feed date cell -> ISO YYYY-MM-DD string or None (never junk)."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # BLN corrections first: they key off the raw text, including the
    # "YYYY-MM-DD 00:00:00" form real datetime cells stringify to.
    if s in _DATE_CORRECTIONS:
        return _DATE_CORRECTIONS[s]
    if isinstance(val, (datetime, date)):
        return _validate_iso(val.strftime("%Y-%m-%d"))
    for fmt in _DATE_FORMATS:
        try:
            return _validate_iso(datetime.strptime(s, fmt).strftime("%Y-%m-%d"))
        except ValueError:
            continue
    return _validate_iso(warn_monitor._safe_date(s))


def _clean_employees(val) -> int:
    """Raw 'Emp #' cell -> int (0 when the state publishes no count)."""
    n = warn_monitor._safe_int(val)
    if n is None or n < 0 or n > _MAX_JOBS:
        return 0
    return n


def _sheet_rows(grid):
    """One sheet's rows -> list of crosswalk-keyed dicts.

    The header row is the one containing both "Company" and "Notice Date";
    everything above it (the "Iowa WARN Log" title row) is skipped. Sheets
    with no such row — the duplicate program-year "Sheet2" compilation —
    yield nothing, which is what keeps its rows from being counted twice.
    Repeated header echoes and company-less rows (grand-total footers) are
    dropped.
    """
    header_idx = None
    headers = []
    for i, row in enumerate(grid):
        cells = [_clean_str(c) for c in row]
        if "Company" in cells and "Notice Date" in cells:
            header_idx = i
            headers = [_CROSSWALK.get(c) for c in cells]
            break
    if header_idx is None:
        return []

    rows = []
    for row in grid[header_idx + 1:]:
        rec = {}
        for field, val in zip(headers, row):
            if field is None or val is None:
                continue
            if isinstance(val, (datetime, date)):
                val = str(val)
            rec[field] = val
        company = _clean_str(rec.get("company"))
        if not company or company in ("Company", "Iowa WARN Log"):
            continue  # footer/total row or header echo
        rows.append(rec)
    return rows


def _extract_workbook_rows(xlsx_source):
    """Every sheet of an XLSX (path or file-like) -> crosswalked dicts."""
    wb = load_workbook(filename=xlsx_source, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        rows.extend(_sheet_rows([list(r) for r in ws.iter_rows(values_only=True)]))
    return rows


class IowaWorkforceDevelopment(Source):
    code = "ia"
    name = "Iowa"
    agency = "Iowa Workforce Development"
    source_url = HOSTPAGE
    cadence = "weekly"

    def _get(self, url, tries=3):
        """Polite GET: browser UA, 60s timeout, up to 3 backed-off tries."""
        last_err = None
        for attempt in range(tries):
            if attempt:
                time.sleep(2 * attempt)
            try:
                resp = requests.get(
                    url, headers={"User-Agent": USER_AGENT}, timeout=60
                )
                resp.raise_for_status()
                return resp
            except requests.RequestException as e:
                last_err = e
                log.warning(f"[IA] GET {url} failed (try {attempt + 1}): {e}")
        raise last_err

    def _discover(self, html):
        """Find the WARN Log workbook link on the host page.

        Vendored from BLN warn-scraper ia.py: the anchor whose text
        mentions "Excel" ("WARN Log (Microsoft Excel File)").
        """
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a"):
            text = a.get_text(" ", strip=True)
            href = a.get("href") or ""
            if "excel" in text.lower() and href:
                return urljoin(BASEURL, href)
        raise RuntimeError(
            "IA WARN Log Excel link not found on host page "
            "(layout changed or feed moved)"
        )

    def fetch(self, force: bool = False) -> tuple:
        self.paths.ensure()
        host = self._get(HOSTPAGE)
        current_url = self._discover(host.text)

        time.sleep(1)  # politeness: max 1 request/second/host
        current = self._get(current_url)
        rows = _extract_workbook_rows(io.BytesIO(current.content))

        # 2011-2018 backfill from BLN's public archive (third-party bucket:
        # tolerate an outage rather than fail the whole state).
        historic_rows = 0
        try:
            time.sleep(1)
            historic = self._get(HISTORIC_URL)
            hist = _extract_workbook_rows(io.BytesIO(historic.content))
            historic_rows = len(hist)
            rows.extend(hist)
        except requests.RequestException as e:
            log.warning(f"[IA] historic backfill unavailable: {e}")

        self.paths.raw.write_text(json.dumps(rows, indent=1, default=str))

        meta = warn_monitor._load_meta(self.paths.meta)
        new_hash = warn_monitor._file_hash(self.paths.raw)
        changed = force or new_hash != meta.get("file_hash", "")
        meta.update(
            {
                "file_hash": new_hash,
                "last_checked": datetime.now(timezone.utc).isoformat(),
                "url": current_url,
                "historic_url": HISTORIC_URL,
                "historic_rows": historic_rows,
                "row_count": len(rows),
            }
        )
        warn_monitor._save_meta(meta, self.paths.meta)
        return changed, str(self.paths.raw)

    def parse(self, raw_path) -> pd.DataFrame:
        rows = json.loads(Path(raw_path).read_text())
        out = []
        for r in rows:
            company = _clean_str(r.get("company"))
            if not company:
                continue
            out.append(
                {
                    "company": company,
                    "notice_date": _clean_date(r.get("notice_date")),
                    "effective_date": _clean_date(r.get("effective_date")),
                    "employees": _clean_employees(r.get("employees")),
                    "layoff_type": _clean_str(r.get("layoff_type")),
                    "county": _clean_str(r.get("county")),
                    "city": _clean_str(r.get("city")),
                    "address": _clean_str(r.get("address")),
                    "industry": _clean_str(r.get("industry")),
                }
            )
        return pd.DataFrame(out, columns=_OUTPUT_COLUMNS)
