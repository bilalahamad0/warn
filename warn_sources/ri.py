"""
warn_sources.ri
---------------
Rhode Island — Department of Labor and Training WARN Report workbook.

The DLT WARN page links one multi-sheet XLSX ("WARN Report") with a sheet
per recent year plus "Previous Years" (data back to 2010). The file's URL
moves whenever DLT uploads a fresh copy, so ``fetch`` re-discovers the link
from the landing page, then delegates to ``warn_monitor.download_xlsx`` for
conditional-GET caching against the resolved URL.

Parse/crosswalk logic vendored from Big Local News' Apache-2.0 warn-scraper
(warn/scrapers/ri.py) and warn-transformer (transformers/ri.py).

Feed quirks honored (all observed live or in BLN's corrections):
- a title row + per-sheet header row precede the data on every sheet;
- the company header oscillates between "Company Name " and
  "Company Name (* Denotes Covid 19 Related WARN)" — "*" marks are noise;
- dates are often ranges/lists ("12/8/2024 through 12/21/2024",
  "5/2/24-7/1/24", "10/21/23 & 12/30/23") — keep the first date;
- typo years ("2108-10-23", "5/4/204") and "Staggered" appear;
- "Number Affected" carries prose ("54 Union 3 Non Union", "60-80",
  "9,891 Remote Workers (2 from RI)") needing manual corrections.
"""

import dataclasses
import logging
import re
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import warn_monitor
from .base import Source, StatePaths

log = logging.getLogger("warn_sources.ri")

_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)

_ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# Manual corrections vendored from BLN warn-transformer transformers/ri.py.
# Date keys are str(cell) — openpyxl yields datetime for the "2108" typo
# cells, whose str() form is "%Y-%m-%d %H:%M:%S".
_DATE_CORRECTIONS = {
    "2108-10-23 00:00:00": "2018-10-23",
    "2108-11-01 00:00:00": "2018-11-01",
    "Staggered": None,
}

# Jobs keys are str(cell) stripped with commas removed (BLN transform_jobs
# normalization). The "1900-*" keys are Excel date-cell accidents.
_JOBS_CORRECTIONS = {
    "---": None,
    "54 Union 3 Non Union": 57,
    "190 company with an additional 100 contracted": 290,
    "60-80": 60,
    "additional 16": 16,
    "1900-03-17 00:00:00": None,
    "309 *updated 10/26/23": 309,
    "1900-01-01 00:00:00": 1,
    "16 (additional)": 16,
    "1900-02-07 00:00:00": 38,
    "9891 Remote Workers (2 from RI)": 2,
    "1 (Remote worker)": 1,
}

_MAX_JOBS = 10_000  # BLN maximum_jobs sanity cap


def _iso_or_none(val):
    """_safe_date, but strictly ISO with a plausible year — else None."""
    s = warn_monitor._safe_date(val)
    if not s or not _ISO_RE.match(s):
        return None
    year = int(s[:4])
    if year < 1988 or year > date.today().year + 2:  # "5/4/204", "2108-…"
        return None
    return s


def _ri_date(val):
    """WARN/Effective date cell -> ISO YYYY-MM-DD or None."""
    if val is None:
        return None
    key = str(val).strip()
    if not key:
        return None
    if key in _DATE_CORRECTIONS:
        return _DATE_CORRECTIONS[key]
    iso = _iso_or_none(val)
    if iso is None and not isinstance(val, (datetime, date)):
        # BLN fallback for ranges/lists: keep the first date token.
        cleaned = key.split()[0].split("-")[0]
        cleaned = cleaned.replace("–", "").replace(",", "").strip()
        iso = _iso_or_none(cleaned) if cleaned else None
    return iso


def _ri_jobs(val) -> int:
    """'Number Affected' cell -> int (0 when no usable count published)."""
    if val is None:
        return 0
    key = str(val).strip().replace(",", "")
    if key in _JOBS_CORRECTIONS:
        n = _JOBS_CORRECTIONS[key]
    else:
        n = warn_monitor._safe_int(val)
    if n is None or n < 0 or n > _MAX_JOBS:
        return 0
    return int(n)


def _ri_text(val) -> str:
    """Whitespace-collapsed text cell ('' when empty)."""
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


def _ri_company(val) -> str:
    """Company cell, minus the '* Denotes Covid 19 Related' markers."""
    return _ri_text(str(val).replace("*", ""))


def _ri_layoff_type(val) -> str:
    """'Closing Yes/No' -> Closure / Layoff / '' (blank when unstated)."""
    s = str(val).strip().lower() if val is not None else ""
    if "yes" in s:  # BLN check_if_closure quirk: any 'yes' marks a closure
        return "Closure"
    if s.startswith("no"):
        return "Layoff"
    return ""


def _norm_header(val) -> str:
    """Normalize a header cell; company-name variants collapse to one."""
    h = str(val).strip() if val is not None else ""
    if h.startswith("Company Name"):
        return "Company Name"
    return h


class RhodeIslandDLT(Source):
    code = "ri"
    name = "Rhode Island"
    agency = "Rhode Island Department of Labor and Training"
    source_url = (
        "https://dlt.ri.gov/employers/"
        "worker-adjustment-and-retraining-notification-warn"
    )
    cadence = "twice-daily"

    def make_paths(self, data_dir: Optional[Path] = None) -> StatePaths:
        # Standard layout, but openpyxl refuses extensionless files.
        paths = super().make_paths(data_dir)
        return dataclasses.replace(paths, raw=paths.raw.with_suffix(".xlsx"))

    def _discover_xlsx_url(self) -> str:
        """Find the current WARN Report XLSX link on the landing page."""
        last_err = None
        for attempt in range(3):
            if attempt:
                time.sleep(2 * attempt)
            try:
                resp = requests.get(
                    self.source_url, headers={"User-Agent": _UA}, timeout=60
                )
                resp.raise_for_status()
                break
            except requests.RequestException as e:
                log.warning(f"RI landing page attempt {attempt + 1} failed: {e}")
                last_err = e
        else:
            raise RuntimeError(f"RI: landing page fetch failed: {last_err}")

        soup = BeautifulSoup(resp.text, "html.parser")
        # Primary: the link labeled "WARN Report" (BLN's selector).
        for link in soup.find_all("a"):
            href = link.get("href") or ""
            if href and "WARN Report" in link.get_text():
                return urljoin(self.source_url, href)
        # Fallback: any Excel link whose path mentions WARN.
        for link in soup.find_all("a"):
            href = (link.get("href") or "").split("?")[0]
            if "warn" in href.lower() and href.lower().endswith((".xlsx", ".xls")):
                return urljoin(self.source_url, link.get("href"))
        raise RuntimeError(
            f"RI: no 'WARN Report' XLSX link found on {self.source_url}"
        )

    def fetch(self, force: bool = False) -> tuple:
        excel_url = self._discover_xlsx_url()
        time.sleep(1)  # politeness: max 1 request/second/host
        if not self.paths.raw.exists():
            force = True  # a 304 is useless when the cached file is gone
        return warn_monitor.download_xlsx(
            force=force,
            url=excel_url,
            meta_file=self.paths.meta,
            local_path=self.paths.raw,
        )

    def parse(self, raw_path) -> pd.DataFrame:
        wb = load_workbook(filename=str(raw_path), read_only=True, data_only=True)
        records = []
        try:
            for ws in wb.worksheets:
                cols = None
                for row in ws.iter_rows(values_only=True):
                    if not any(
                        v is not None and str(v).strip() for v in row
                    ):
                        continue
                    headers = [_norm_header(v) for v in row]
                    if "WARN Date" in headers and "Company Name" in headers:
                        cols = {h: i for i, h in enumerate(headers) if h}
                        continue
                    if cols is None:
                        continue  # title/preamble rows above the header
                    rec = self._row_to_record(row, cols)
                    if rec is not None:
                        records.append(rec)
        finally:
            wb.close()
        if not records:
            raise ValueError(f"RI: no WARN rows parsed from {raw_path}")
        df = pd.DataFrame(records)
        for col in ("notice_date", "effective_date"):
            # DataFrame construction turns None into NaN; keep the
            # contract literal: ISO string or None.
            df[col] = df[col].astype(object).where(df[col].notna(), None)
        return df

    @staticmethod
    def _row_to_record(row, cols):
        def cell(name):
            i = cols.get(name)
            return row[i] if i is not None and i < len(row) else None

        raw_company = cell("Company Name")
        if raw_company is None or "Company Name" in str(raw_company):
            return None  # blank line or a repeated header row
        company = _ri_company(raw_company)
        if not company or company == "Rhode Island WARN Report":
            return None
        return {
            "company": company,
            "notice_date": _ri_date(cell("WARN Date")),
            "effective_date": _ri_date(cell("Effective Date")),
            "employees": _ri_jobs(cell("Number Affected")),
            "layoff_type": _ri_layoff_type(cell("Closing Yes/No")),
            "city": _ri_text(cell("Location of Layoffs")),
        }
