"""
warn_sources.nj
---------------
New Jersey — Department of Labor and Workforce Development WARN archive.

Easiest big-state feed: the entire 2004-present archive ships as ONE XLSX
(`WARN_Notice_Archive.xlsx`, one worksheet per year) with proper ETag /
Last-Modified caching, so fetch is a straight ``warn_monitor.download_xlsx``.

Each sheet has exactly five columns: Company, City, Month Posted,
Effective Date, Workforce Affected. NJ publishes no true notice date —
"Month Posted" is the month the state posted the notice, not a notice date —
so ``notice_date`` stays None, never synthesized (EXPANSION_RESEARCH.md §5).

Parse/crosswalk logic vendored from Big Local News' Apache-2.0 warn-scraper
(warn/scrapers/nj.py) and warn-transformer (transformers/nj.py).

Known feed quirks handled here (all observed in the live workbook):
- Effective Date cells are mostly real datetimes but ~10% are strings:
  single dates ("3/4/07"), ranges ("6/13/25 - 8/22/25"), multi-date lists
  ("2/13/24,3/15/24"), prose ("Rolling basis beginning on 6/4/25"), and
  placeholders ("TBA", "TBD", "-", "Temp layoff"). Strategy: BLN's
  hand-audited corrections first, then strict single-date formats, then the
  first date-like token (BLN's own rule for nearly every range/list).
- Workforce Affected mixes ints with "*149"-style asterisked strings
  (rehired-by-buyer footnote marker), "TBA"/"To be Determined"/"-", and one
  multi-county breakdown that BLN sums to 871.
- Junk rows: repeated blank rows, a footnote row in the 2009 sheet, and one
  orphan row (a stray date with no company) — all dropped.
"""

import dataclasses
import re
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

import warn_monitor
from .base import Source, StatePaths

# Placeholders NJ uses for "no date yet".
_NON_DATES = {"", "-", "TBA", "TBD", "Temp layoff", "Unknown", "N/A"}

_DATE_FORMATS = ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d")
_DATE_TOKEN = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")

# Vendored from BLN warn-transformer nj.py ``date_corrections``: only the
# entries where BLN's hand-audited pick differs from the first-date-token
# rule (typos, out-of-order lists, unsalvageable multi-notice strings).
_DATE_CORRECTIONS = {
    "3030-08-23 00:00:00": "2020-08-23",
    "08/15/2023, 8/22/2023": "2024-08-15",
    "4/5/2024, 3/31/2024": "2024-03-31",
    "4/5/24, 3/31/24": "2024-03-31",
    "11/7/25,12/26/25,3/28/25": "2025-03-28",
    "4/24/25, 7/3/25, 7/18/25, 7/31/25, 12/11/25": "2024-04-25",
    "9/6/204": "2024-09-06",
    "7/723 -  08/23": "2023-07-07",
    "7/723 -  8/2023": "2023-07-07",
    (
        "12/31/24, 9/27/24, 8/30/24, 5/31/24, 1/31/24, 12/13/2023, "
        "10/25/2023, 9/27/2023, 9/20/2023, 9/18/2023,9/7/2023, "
        "4/13/23, 3/30/23"
    ): None,
    (
        "12/31/2024, 9/27/2024, 8/30/2024, 5/31/2024, 1/31/2024, "
        "12/13/2023, 10/25/2023, 9/27/2023, 9/20/2023, 9/18/2023,"
        "9/7/2023, 4/13/2023, 3/30/2023"
    ): None,
}

# Vendored from BLN warn-transformer nj.py ``jobs_corrections`` (the 16000
# United Airlines figure is legitimate, though nationwide; 23695 is a
# feed error on a 2020 usi services group row). The comma variant of the
# multi-county Amazon breakdown is what the live workbook actually holds.
_JOBS_CORRECTIONS = {
    "TBA": None,
    "TBD": None,
    "To be Determined": None,
    "-": None,
    "Unknown": None,
    "": None,
    23695: None,
    "23695": None,
    16000: 16000,
    "240 (Passaic) 417 (Bergen) 141 (Monmouth) 44 (Hudson) 29 (Statewide)": 871,
    (
        "240 (Passaic), 417 (Bergen), 141 (Monmouth), "
        "44 (Hudson), 29 (Statewide)"
    ): 871,
    "80 - 100": 80,
}


def _iso(dt) -> Optional[str]:
    """ISO date with a sanity window (guards '9/6/204'-style year typos)."""
    if not 1988 <= dt.year <= 2040:
        return None
    return dt.strftime("%Y-%m-%d")


def _effective_date(val) -> Optional[str]:
    """NJ Effective Date cell -> ISO YYYY-MM-DD string or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return _iso(val)
    s = str(val).strip()
    if s in _NON_DATES:
        return None
    if s in _DATE_CORRECTIONS:
        return _DATE_CORRECTIONS[s]
    for fmt in _DATE_FORMATS:
        try:
            return _iso(datetime.strptime(s, fmt))
        except ValueError:
            pass
    # Ranges / multi-date lists / prose: BLN's rule is the first date.
    m = _DATE_TOKEN.search(s)
    if m:
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return _iso(datetime.strptime(m.group(0), fmt))
            except ValueError:
                pass
    return None


def _employees(val) -> int:
    """NJ Workforce Affected cell -> int (0 when the state published none)."""
    if val is None:
        return 0
    if val in _JOBS_CORRECTIONS:
        fixed = _JOBS_CORRECTIONS[val]
        return fixed if fixed is not None else 0
    if isinstance(val, str):
        s = val.strip()
        if s in _JOBS_CORRECTIONS:
            fixed = _JOBS_CORRECTIONS[s]
            return fixed if fixed is not None else 0
        # BLN transform_jobs: cut the asterisk they sometimes use ("*149").
        val = s.replace("*", "")
    n = warn_monitor._safe_int(val)
    return n if n is not None else 0


class NewJerseyDOL(Source):
    code = "nj"
    name = "New Jersey"
    agency = "New Jersey Department of Labor and Workforce Development"
    source_url = (
        "https://www.nj.gov/labor/assets/PDFs/WARN/WARN_Notice_Archive.xlsx"
    )
    cadence = "monthly"

    def make_paths(self, data_dir: Optional[Path] = None) -> StatePaths:
        # openpyxl validates by file extension, so the raw download needs
        # a real .xlsx suffix (the standard layout's ``raw_download`` has
        # none).
        paths = super().make_paths(data_dir)
        return dataclasses.replace(paths, raw=paths.root / "raw_download.xlsx")

    def fetch(self, force: bool = False) -> tuple:
        return warn_monitor.download_xlsx(
            force=force,
            url=self.source_url,
            meta_file=self.paths.meta,
            local_path=self.paths.raw,
        )

    def parse(self, raw_path) -> pd.DataFrame:
        # Row loop vendored from BLN warn-scraper nj.py: every worksheet
        # (one per year), positional columns, skip each sheet's header row.
        rows = []
        with warnings.catch_warnings():
            # The state's workbook carries an unparseable print header;
            # read-only sheets parse lazily, so the filter must span the
            # whole row loop.
            warnings.filterwarnings("ignore", message="Cannot parse header")
            wb = load_workbook(filename=raw_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    vals = list(row[:5]) + [None] * max(0, 5 - len(row))
                    company, city, _month_posted, eff, jobs = vals
                    company = str(company).strip() if company is not None else ""
                    city = str(city).strip() if city is not None else ""
                    # Drop repeated headers, blank, footnote/orphan rows.
                    if not company or company.lower() == "company":
                        continue
                    if not city and eff is None and jobs is None:
                        continue
                    rows.append(
                        {
                            "company": company,
                            "city": city,
                            "effective_date": _effective_date(eff),
                            "employees": _employees(jobs),
                        }
                    )
            wb.close()
        return pd.DataFrame(
            rows, columns=["company", "city", "effective_date", "employees"]
        )
